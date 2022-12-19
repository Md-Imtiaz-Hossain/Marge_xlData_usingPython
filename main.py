import xlwings as xw
import openpyxl as xl
import pandas as pd
class information:

    def __init__(self, database):
        self.database = xw.Book(database)

    def get_dataset_column(self, sheet, data_range):
        return self.database.sheets[sheet].range(data_range).value


source_database_path = input("Please input path for input: ")
# output_database_path = input("Please input path for output: ")
source_database = information(source_database_path)
# DATASET/Details Data Final Fall-2007.xls
# output_database = xl.load_workbook(output_database_path)["Result 1"]
# Database Created
output_database = pd.DataFrame()
# for _ in range(1): #Number of sheets
current_sheet_name = input("Please input sheet name: ")
how_many_data = input("How many columns will you select from this sheet?: ")
for i in range(int(how_many_data)): #Get info of specific column
    data_name = input("Please input what data you are selecting: ")
    range_data = input("Please input range for " + data_name + ": ")
    data = source_database.get_dataset_column(current_sheet_name, range_data)
    output_database.insert(i, data_name, data)
    print("Done inserting " + data_name + " column, ready for next...")


output_database.to_csv('pandas/' + current_sheet_name + ".csv", index= False)
print(output_database)




# Dataset path
# DATASET/23.  Final Report Print -(All Information) - Statistic Spring-2017.xlsx
# DATASET/Details Data (correction) 2006 Spring.xls
# DATASET/Details Data Final - Fall-  2013 Final.xls
# DATASET/Master_Details_Data Fall 2015 09.01.15 Statistics.xls


# from openpyxl import load_workbook

#
# ws = xw.Book("DATASET/Details Data Final Fall-2007.xls").sheets['CSE']
# #ws1 = xl.Book("F:\Work\TABLE TEMPLATE\ADMISSION_STUDENTS.xlsx")
# XL = xl.load_workbook("TABLE TEMPLATE/ADMISSION_STUDENTS.xlsx")
# sheet1 = XL["Result 1"]
# #
# ###### Father and Mother #############
#
# v1 = ws.range("F8:F64").value
# v1_name = ws.range("D8:D64").value
# print(v1_name)
# v3 = len(v1)
# v2 = []
# print(v3)
# for i in range(0, v3):
#     if v1[i] is not None:
#         v2.append(v1[i].split(','))
#
# print(v2)
# Father = []
# Mother = []
# v3 = sheet1.max_row
#
# for i in range(0, len(v2)):
#     for j in range(0, 2):
#         if j == 0:
#             Father.append(v2[i][j])
#         if j == 1:
#             Mother.append(v2[i][j])
#
#
# count = 0
# j = 1
# for i in range(0, len(v1_name)):
#     if v1_name[i] is not None:
#         sheet1.cell(row=v3+j, column=16).value = Father[count]
#         count += 1
#         j += 1
#     elif v1_name[i] is None:
#         j += 1
#         continue
# j = 1
# count1 = 0
# for i in range(0, len(v1_name)):
#     if v1_name[i] is not None:
#         sheet1.cell(row=v3+j, column=17).value = Mother[count1]
#         count1 += 1
#         j += 1
#     elif v1_name[i] is None:
#         j += 1
#         continue
#
# v3_new = v3
# ################### END ######################
#
#
#
#
#
# #####  FOR NAME #######
# v1 = ws.range("D8:D64").value
# print("Result:", v1)
#
# v2 = len(v1)
# v3 = v3_new
#
# print(v3, v2)
# count = 0
# for i in range(v3, v3+v2):
#     sheet1.cell(row=i+1, column=15).value = v1[count]
#     count += 1
#
# v3_new = v3
# ###### NAME END #######
#
#
#
# #####  FOR STUDENT_ID #######
# v1 = ws.range("B8:B64").value
# print("Result:", v1)
#
# v2 = len(v1)
# v3 = v3_new
#
#
# print(v3, v2)
# count = 0
# for i in range(v3, v3+v2):
#     sheet1.cell(row=i+1, column=23).value = v1[count]
#     count += 1
# v3_new = v3
# #####  END STUDENT_ID #######
#
#
# #####  FOR Gender #######
# v1 = ws.range("E8:E64").value
# print("Result:", v1)
#
# v2 = len(v1)
# v3 = v3_new
#
# print(v3, v2)
# count = 0
# for i in range(v3, v3+v2):
#     sheet1.cell(row=i+1, column=13).value = v1[count]
#     count += 1
# v3_new = v3
# #####  END Gender #######
#
#
# #####  FOR SSC_YEAR #######
# v1 = ws.range("O8:O64").value
# print("Result:", v1)
#
# v2 = len(v1)
# v3 = v3_new
#
# print(v3, v2)
# count = 0
# for i in range(v3, v3+v2):
#     sheet1.cell(row=i+1, column=11).value = v1[count]
#     count += 1
# v3_new = v3
# #####  END SSC_YEAR #######
#
#
#
#
# #####  FOR SSC_BOARD #######
# v1 = ws.range("J8:J64").value
# print("Result:", v1)
#
# v2 = len(v1)
# v3 = v3_new
#
# print(v3, v2)
# count = 0
# for i in range(v3, v3+v2):
#     sheet1.cell(row=i+1, column=9).value = v1[count]
#     count += 1
# v3_new = v3
# #####  END SSC_BOARD #######
#
#
#
# #####  FOR HSC_YEAR #######
# v1 = ws.range("P8:P64").value
# print("Result:", v1)
#
# v2 = len(v1)
# v3 = v3_new
#
# print(v3, v2)
# count = 0
# for i in range(v3, v3+v2):
#     sheet1.cell(row=i+1, column=7).value = v1[count]
#     count += 1
# v3_new = v3
# #####  END HSC_YEAR #######
#
#
# #####  FOR HSC_BOARD #######
# v1 = ws.range("K8:K64").value
# print("Result:", v1)
#
# v2 = len(v1)
# v3 = v3_new
#
# print(v3, v2)
# count = 0
# for i in range(v3, v3+v2):
#     sheet1.cell(row=i+1, column=4).value = v1[count]
#     count += 1
# v3_new = v3
# #####  END hSC_BOARD #######
#
#
#
# #####  FOR SSC_GPA #######
# v1 = ws.range("L8:L64").value
# print("Result:", v1)
#
# v2 = len(v1)
# v3 = v3_new
#
# print(v3, v2)
# count = 0
# for i in range(v3, v3+v2):
#     sheet1.cell(row=i+1, column=18).value = v1[count]
#     count += 1
# v3_new = v3
# #####  END SSC_GPA #######
#
#
#
# #####  FOR HSC_GPA #######
# v1 = ws.range("M8:M64").value
# print("Result:", v1)
#
# v2 = len(v1)
# v3 = v3_new
#
# print(v3, v2)
# count = 0
# for i in range(v3, v3+v2):
#     sheet1.cell(row=i+1, column=19).value = v1[count]
#     count += 1
# v3_new = v3
# #####  END HSC_GPA #######
#
#
# #####  FOR RELIGION #######
# v1 = ws.range("I8:I64").value
# print("Result:", v1)
#
# v2 = len(v1)
# v3 = v3_new
#
# print(v3, v2)
# count = 0
# for i in range(v3, v3+v2):
#     sheet1.cell(row=i+1, column=46).value = v1[count]
#     count += 1
# v3_new = v3
# #####  END RELIGION #######
#
#
# #####  FOR SEMESTER ID  #######
#
# v1 = ws.range("I8:I64").value
# print("Result:", v1)
#
# v2 = len(v1)
# v3 = v3_new
#
# print(v3, v2)
# #count = 0
# for i in range(v3, v3+v2):
#     sheet1.cell(row=i+1, column=1).value = "11022007"
#     #count += 1
# v3_new = v3
#
# #####  END SEMESTER ID  #######
#
# XL.save('TABLE TEMPLATE/ADMISSION_STUDENTS-new.xlsx')
#
